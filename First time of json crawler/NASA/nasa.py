import os

import requests
from loguru import logger
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from datetime import datetime


def get_page_data(page_number):
    url = "https://osdr.nasa.gov/geode-py/ws/repo/search"
    params = {
        "term": "kidney",
        "source": "cgene,alsda",
        "type": "study",
        "from": str(page_number),
        "size": "100",
        "sort": "Study Public Release Date",
        "order": "desc"
    }
    response = requests.get(url, params=params).json()
    hits = response['hits']['hits']
    for hit in hits:
        Title = hit['_source']['Study Title']
        Identifier = hit['_source']['Accession']
        Organisms = hit['_source']['organism']
        Factors = hit['_source']['Study Factor Name']
        assayTypes = hit['_source']['Study Assay Measurement Type']
        releaseDate = int(hit['_source']['Study Public Release Date'])
        timeDate = datetime.fromtimestamp(releaseDate).strftime('%Y-%m-%d')
        Description = hit['_source']['Study Description']
        study_link = f"https://osdr.nasa.gov/bio/repo/data/studies/{Identifier}"
        write_to_excel([Title, Identifier, Organisms, Factors, assayTypes, timeDate, Description, study_link], 'nasa')
        logger.info(Title, Identifier, Organisms, Factors, assayTypes, timeDate, Description, study_link)


def write_to_excel(val, title):
    file_path = os.path.join(os.getcwd(), f'{title}.xlsx')
    if os.path.exists(file_path) is False:
        wb = Workbook()
        wb.create_sheet(title, 0)
        sheet = wb[title]
        sheet.append(['Title', 'Identifier', 'Organisms', 'Factors', 'assayTypes', 'timeDate', 'Description','Link'])
        wb.save(file_path)
    wb = load_workbook(file_path)
    sheet = wb[title]
    sheet._current_row = sheet.max_row
    sheet.append(val)
    wb.save(file_path)


if __name__ == '__main__':
    get_page_data(0)
